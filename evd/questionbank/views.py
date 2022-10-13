from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from .forms import QuestionForm, get_question_data
from .models import *
from web.models import *
from web.views import has_role, has_pref_role
from django.db.models import Q
import json
import simplejson
import re
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from openpyxl import load_workbook
import genutilities.views as genUtility
import genutilities.logUtility as logService
import settings
import traceback
@login_required
def addQuestion(request):
    if request.method == "GET":
        is_eligible = True
        question_form = QuestionForm()
        # if has_role(request.user.userprofile,"Well Wisher"):
        #     return HttpResponse
        userprofile_id =  UserProfile.objects.values_list("id",flat = True).filter(user_id = request.user.id)
        role_id = RolePreference.objects.filter(userprofile_id = userprofile_id,role_id =3,role_status = "Active",role_outcome = "Recommended")
        if request.user.is_superuser or role_id:
            boards = Center.objects.filter(status="Active").values_list("board", flat=True).distinct().order_by("board")
        elif  has_role(request.user.userprofile, "Teacher") or has_pref_role(request.user.userprofile, "Teacher"):

            boards= Offering.objects.filter(active_teacher__id=request.user.id).values_list("center__board",flat=True).distinct()
            if len(boards) == 0:
                is_eligible = False
        else:
            boards = Center.objects.filter(Q(admin=request.user) | Q(assistant=request.user) 
                                       | Q(field_coordinator_id=request.user.id) | Q(delivery_coordinator_id=request.user.id)
                                       ).filter(status="Active").values_list("board", flat=True).distinct().order_by("board")
        if len(boards) == 0:
            try:
                boards = Center.objects.filter(Q(funding_partner_id=request.user.partner_set.all()[0].id) \
                                       | Q(delivery_partner=request.user.partner_set.all()[0].id) | Q(orgunit_partner=request.user.partner_set.all()[0].id) \
                                       ).filter(status="Active").values_list("board", flat=True).distinct().order_by("board")
            except IndexError:
                boards = ''
        board_choices = [("","Select Board")]+[(i, i) for i in boards]
        board_choices = tuple(board_choices)
        question_form.fields["board"].choices = board_choices
        return render(request, 'add_or_update_question.html', {"is_eligible": is_eligible, 'question_form': question_form})
    elif request.method == "POST":
        user_action_type = request.POST.get('action_type')
        board = request.POST.get('board')
        subject = request.POST.get('subject')
        grade = request.POST.get('grade')
        main_topic = request.POST.get('topic')  # Id value will be received, not the string
        subtopic = request.POST.get('subTopic', "")
        exercise_number = request.POST.get('exerciseNumber')
        question_type = request.POST.get('questionType')
        print "Question Type ", question_type
        difficulty_level = request.POST.get('difficultyLevel')
        question_instructions = request.POST.get('question_instruction', "")
        main_question = request.POST.get('question')
        delete_image = request.POST.get('delete_image',"")
        uploaded_image_href = request.POST.get('saved_href',"")
        uploaded_image = request.FILES.get('questionImage',"")
        expected_answer = request.POST.get('answer')
        solution = request.POST.get('solution', "")
        mcq_choice_1 = request.POST.get('choice_1', "")
        mcq_choice_2 = request.POST.get('choice_2', "")
        mcq_choice_3 = request.POST.get('choice_3', "")
        mcq_choice_4 = request.POST.get('choice_4', "")
        mcq_choice_5 = request.POST.get('choice_5', "")
        print("$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$",delete_image,uploaded_image_href,uploaded_image)
        if uploaded_image_href == "undefined":
            uploaded_image_href = ""
        if mcq_choice_4 == "undefined":
            mcq_choice_4 = ""
        if mcq_choice_5 == "undefined":
            mcq_choice_5 = ""
        expected_mcq_answer = request.POST.get('mcqAnswer', "")
        expected_true_false_answer = request.POST.get('trueFalseAnswer', "")
        marks = request.POST.get('marks')
        quesId = request.POST.get('id')
        saveAndAdd = request.POST.get('saveAndAdd')
        answer_question = ''
        correct_answer = ''

        response_message = "Question already exists!"
        redirect_url = ""
        if user_action_type == 'update':
            print '*****running*****',quesId
            if question_type == "MCQs" or question_type == "Multiple Choices":
                answer_question = ([mcq_choice_1, mcq_choice_2, mcq_choice_3, mcq_choice_4, mcq_choice_5])
                correct_answer = [expected_mcq_answer,'']
            elif question_type == "DirectAnswer" or question_type == "Direct Answer" or question_type == "Others":
                correct_answer = expected_answer
            elif question_type == "True/False":
                answer_question = [u'True',u'False']
                correct_answer = expected_true_false_answer
            else:
                correct_answer = expected_answer
            if subtopic == 'null':
                subtopic = None
            newPhoto = ''
            if not uploaded_image and delete_image == "false":
                newPhoto = uploaded_image_href
            elif uploaded_image:
                image_name =uploaded_image.name.split(':')
                for pht in image_name:
                        newPhoto = pht
                f_path = os.getcwd()
                f_name = '/static/questionbank/question/images/' + newPhoto +'' 
                f = open(f_path + f_name, 'w+')
                f.write(uploaded_image.read())
                f.close()
            else:
                newPhoto=uploaded_image
            Question.objects.filter(id=quesId).update(board = board, subject = subject, grade = grade, topic = main_topic, subtopic = subtopic, question = main_question, question_instruction = question_instructions, question_image = newPhoto, question_type = question_type, difficulty_level = difficulty_level, answer_question = answer_question, correct_answer = correct_answer,solution = solution, marks = marks, status= 'Submitted')
            response_message = "Question is updated successfully."
            redirect_url = "/questions/"
        # Check if question is already added in the database
        else:
            try:
                question_obj = Question.objects.get(board = board, subject = subject, grade = grade, topic = main_topic, question = main_question)
            except Question.DoesNotExist:
                if question_type == "MCQs" or question_type == "Multiple Choices":
                    answer_question = ([mcq_choice_1, mcq_choice_2, mcq_choice_3, mcq_choice_4, mcq_choice_5])
                    correct_answer = [expected_mcq_answer,'']
                elif question_type == "DirectAnswer" or question_type == "Direct Answer" or question_type == "Others":
                    correct_answer = expected_answer
                elif question_type == "True/False":
                    answer_question = [u'True',u'False']
                    correct_answer = expected_true_false_answer
                else:
                    correct_answer = expected_answer
                question_image = ''
                if uploaded_image:
                    image_name =uploaded_image.name.split(':')
                    for pht in image_name:
                            question_image = pht
                    f_path = os.getcwd()
                    f_name = '/static/questionbank/question/images/' + question_image +'' 
                    f = open(f_path + f_name, 'w+')
                    f.write(uploaded_image.read())
                    f.close()
                question_obj = Question.objects.create(board = board, subject = subject, grade = grade, topic_id = main_topic, exercise_number = exercise_number,
                    subtopic_id = subtopic, question = main_question, question_instruction = question_instructions,
                    question_image = question_image, question_type = question_type, difficulty_level = difficulty_level,
                    answer_question = answer_question, correct_answer = correct_answer,solution = solution,marks = marks,
                    added_by = request.user, status= 'Submitted')


                question_obj.save()

                response_message = "Question is saved successfully."
                if saveAndAdd == 'saveadd':
                    print "saveandAdd",saveAndAdd
                    redirect_url = "/questions/add/"
                else:
                    print "Save",saveAndAdd
                    redirect_url = "/questions/"

        return HttpResponse(simplejson.dumps({'message': response_message, 'redirect_url': redirect_url}),
            mimetype = 'application/json')


@login_required
def view_or_list_Questions(request,question_id=""):
    # who can view all Questions and is there any restriction other than super user
    is_admin = False
    if request.method == "GET":
        userprofile_id =  UserProfile.objects.values_list("id",flat = True).filter(user_id = request.user.id)
        role_id = RolePreference.objects.filter(userprofile_id = userprofile_id,role_id =3,role_status = "Active",role_outcome = "Recommended")
        
        if question_id:
            try:
                question = Question.objects.get(id=question_id)
                isDc = isContentAdmin = isauthor = False
                isContentAdmin = has_role(request.user.userprofile, "Content Admin")
                isDc = has_role(request.user.userprofile, "Delivery co-ordinator")
                isauthor = True if question.added_by == request.user else False
                canEdit = True if (isauthor or request.user.is_superuser or isContentAdmin or isDc) else False
                return render(request,'view_or_list_questions.html',{"question":question,"view_flag":True,"canEdit":canEdit})
            except Question.DoesNotExist:
                error_msg = "Oops !!! Question not Found"
                return render(request,"view_or_list_questions.html",{"err_msg":error_msg,"view_flag":True})
        board_list = ''
        if request.user.is_superuser or role_id:
            is_admin = True
            board_list = Center.objects.filter(status="Active").values_list("board", flat=True).distinct().order_by("board")
        elif  has_role(request.user.userprofile, "Teacher") or has_pref_role(request.user.userprofile, "Teacher"):

            board_list = Offering.objects.filter(active_teacher__id=request.user.id).values_list("center__board",flat=True).distinct()
        else:
            board_list = Center.objects.filter(Q(admin=request.user) | Q(assistant=request.user) 
                                       | Q(field_coordinator_id=request.user.id) | Q(delivery_coordinator_id=request.user.id)
                                       ).filter(status="Active").values_list("board", flat=True).distinct().order_by("board")
        if len(board_list) == 0:
            try:
                board_list = Center.objects.filter(Q(funding_partner_id=request.user.partner_set.all()[0].id) \
                                       | Q(delivery_partner=request.user.partner_set.all()[0].id) | Q(orgunit_partner=request.user.partner_set.all()[0].id) \
                                       ).filter(status="Active").values_list("board", flat=True).distinct().order_by("board")
            except IndexError:
                board_list = ''

        
        questions = Question.objects.exclude(is_deleted=True).filter(board__in = board_list).order_by('-id')
        board_name = request.GET.get('board', '')
        grade = request.GET.get('grade', '')
        subject = request.GET.get('subject', '')
        status = request.GET.get('status', '')
        if board_name and grade and subject and status:
            if board_name != 'All':
                questions = questions.filter(board = board_name)
            if grade != 'All':
                questions = questions.filter(grade= grade)
            if subject != 'All':
                questions = questions.filter(subject= subject)
            if status != 'All':
                questions = questions.filter(status = status)
        for each_question in questions:
            sub_id = each_question.subject
        isDc = isContentAdmin = False
        isContentAdmin = has_role(request.user.userprofile, "Content Admin")
        isDc = has_role(request.user.userprofile, "Delivery co-ordinator")
        canDelete = True if (request.user.is_superuser or isContentAdmin or isDc) else False
        return render(request,'view_or_list_questions.html',{"questions":questions,"canDelete":canDelete, "boards" : board_list, "is_superadmin" : is_admin})
    else:
        return HttpResponseRedirect("/questions")


@login_required
def deleteQuestion(request,question_id):
    # who all are can be able to delete
    # Can user delete other user added questions
    if request.method == "GET" and question_id:
        try:
            question = Question.objects.get(id=question_id)
            isDc = isContentAdmin = isauthor = False
            isContentAdmin = has_role(request.user.userprofile, "Content Admin")
            isDc = has_role(request.user.userprofile, "Delivery co-ordinator")
            isauthor = True if question.added_by == request.user else False
            candelete = True if (isauthor or request.user.is_superuser or isContentAdmin or isDc) else False
            if candelete:
                question.is_deleted = True
                question.delete()
                msg = "Question Deleted Successfully"
            else:
                msg = "Oops..! No permission for you to delete"
        except Question.DoesNotExist:
            msg = "Ooops!!! Question Not Found"
        return HttpResponse(msg)
    else:
        return HttpResponseRedirect("/questions")



def getPrefilledDataQuestionForm(request, data):
    question_form = QuestionForm(initial=data)

    if request.user.is_superuser:
        boards = Center.objects.filter(status="Active").values_list("board", flat=True).distinct().order_by("board")
    elif  has_role(request.user.userprofile, "Teacher") or has_pref_role(request.user.userprofile, "Teacher"):

        boards= Offering.objects.filter(active_teacher__id=request.user.id).values_list("center__board",flat=True).distinct()
        if len(boards) == 0:
            is_eligible = False
    else:
            boards = Center.objects.filter(Q(admin=request.user) | Q(assistant=request.user) \
                                       | Q(field_coordinator=request.user) | Q(delivery_coordinator=request.user)
                                       ).filter(status="Active").values_list("board", flat=True).distinct().order_by("board")
            if len(boards)==0:
                partner_roles=request.user.partner_set.all()[0].id
                if partner_roles :
                    boards=Center.objects.filter(Q(funding_partner=partner_roles) \
                                       | Q(delivery_partner=partner_roles) | Q(orgunit_partner=partner_roles)
                                       ).filter(status="Active").values_list("board", flat=True).distinct().order_by("board")
            if len(boards) == 0:
                is_eligible = False
    board_choices = [("", "Select Board")] + [(i, i) for i in boards]
    question_form.fields["board"].choices = board_choices
    form_board = data["board"]
    subject_choices = getSubjectChoices(form_board)
    question_form.fields["subject"].choices = subject_choices

    form_subject = data["subject"]
    grade_choices = getGradeChoices(form_board, form_subject)
    question_form.fields["grade"].choices = grade_choices
    form_grade = data["grade"]
    topic_choices = getTopicChoices(form_subject,form_board,form_grade)
    question_form.fields["topic"].choices = topic_choices

    form_topic = data["topic"]
    subtopic_choices = getSubTopicChoices(form_topic)
    question_form.fields["subtopic"].choices = subtopic_choices
    return question_form

@login_required
def updateQuestion(request,question_id):
    is_eligible = True
    if request.method == 'GET':
        question_obj = Question.objects.get(id=question_id)
        marks = question_obj.marks
        print("qtype - - -",question_obj.question_type)
        if question_obj.question_type == "MCQs" or question_obj.question_type == "Multiple Choices" or question_obj.question_type == "True/False":
            choices = str(Question.objects.values_list('answer_question').filter(id=question_id))
            ans = str(Question.objects.values_list('correct_answer').filter(id=question_id))
            choices = re.findall(r"'(.*?)'", choices)
            ans = re.findall(r"'(.*?)'", ans)
            for x in choices:
                if ans[-1] == ',' or ans[-1] == '':
                    del ans[-1]

            for x in choices:
                if choices[-1] == ',' or choices[-1] == '':
                    del choices[-1]
            ChoiceLength = len(choices)
            try:
                index_ans = choices.index(ans[0]) + 2
            except:
                index_ans = 4
        else:
            correct_answer = question_obj.correct_answer
            choices = ''
            ChoiceLength = 0
            index_ans = 0
        data = get_question_data(question_obj)
        question_form = getPrefilledDataQuestionForm(request, data)
        return render(request, 'add_or_update_question.html', {'is_eligible':is_eligible,'question_form': question_form, "update_flag":True,"choices":choices,"choiceLength":ChoiceLength,'index_ans':index_ans,'marks':marks})
    elif request.method == "POST":
        question_form = QuestionForm(request.POST)
        if request.user.is_superuser:
            boards = Center.objects.filter(status="Active").values_list("board", flat=True).distinct().order_by("board")
        elif  has_role(request.user.userprofile, "Teacher") or has_pref_role(request.user.userprofile, "Teacher"):

            boards= Offering.objects.filter(active_teacher__id=request.user.id).values_list("center__board",flat=True).distinct()
            if len(boards) == 0:
                is_eligible = False
        else:
            boards = Center.objects.filter(Q(admin=request.user) | Q(assistant=request.user) \
                                       | Q(field_coordinator=request.user) | Q(delivery_coordinator=request.user)
                                       ).filter(status="Active").values_list("board", flat=True).distinct().order_by("board")
            if len(boards)==0:
                partner_roles=request.user.partner_set.all()[0].id
                if partner_roles :
                    boards=Center.objects.filter(Q(funding_partner=partner_roles) \
                                       | Q(delivery_partner=partner_roles) | Q(orgunit_partner=partner_roles)
                                       ).filter(status="Active").values_list("board", flat=True).distinct().order_by("board")
            if len(boards) == 0:
                is_eligible = False
        board_choices = [("","Select Board")]+[(i, i) for i in boards]
        question_form.fields["board"].choices = board_choices
        form_board = request.POST.get("board", "")
        subject_choices = getSubjectChoices(form_board)
        question_form.fields["subject"].choices = subject_choices

        form_subject = request.POST.get("subject", "")
        grade_choices = getGradeChoices(form_board,form_subject)
        question_form.fields["grade"].choices = grade_choices
        form_grade=request.POST.get("grade", "")
        topic_choices = getTopicChoices(form_subject,form_board,form_grade)
        question_form.fields["topic"].choices = topic_choices

        form_topic = request.POST.get("topic", "")
        subtopic_choices = getSubTopicChoices(form_topic)
        question_form.fields["subtopic"].choices = subtopic_choices

        if question_form.is_valid():
            question_obj = Question.objects.get(id=question_id)
            question_obj.board = question_form.cleaned_data['board']
            question_obj.subject = question_form.cleaned_data['subject']
            question_obj.grade = question_form.cleaned_data['grade']
            question_obj.topic_id = question_form.cleaned_data['topic']
            question_obj.subtopic_id = question_form.cleaned_data['subtopic']
            question_obj.exercise_number = question_form.cleaned_data['exercise_number']
            question_obj.question_instruction = question_form.cleaned_data['question_instruction']
            question_obj.question = question_form.cleaned_data['question']
            question_obj.question_image = question_form.cleaned_data['question_image']
            question_obj.question_type = question_form.cleaned_data['question_type']
            question_obj.difficulty_level = question_form.cleaned_data['difficulty_level']
            question_obj.answer_question = question_form.cleaned_data['answer_question']
            question_obj.correct_answer = question_form.cleaned_data['correct_answer']
            question_obj.solution = question_form.cleaned_data['solution']
            question_obj.marks = question_form.cleaned_data['marks']
            question_obj.is_deleted = True if question_form.cleaned_data['is_deleted'] == "True" else False
            question_obj.updated_by = request.user
            question_obj.save()
            return HttpResponseRedirect("/questions/%s/" % question_obj.id)
        else:
            return render(request, 'add_or_update_question.html', {'is_eligible':is_eligible,'question_form': question_form})
    else:
        return HttpResponseRedirect("/questions")


def getSubjectChoices(board):
    subjects = Course.objects.filter(board_name=board).values_list("subject").distinct()
    subject_choices = [(i[0], i[0] ) for i in subjects]
    return subject_choices

def getGradeChoices(board,subject):
    # print "subject",subject
    db_grades = Course.objects.filter(board_name=board, subject=subject).values_list("grade", flat=True).distinct().order_by("grade")
    # print "db_grades",db_grades
    grades_list = []
    for grades in db_grades:
        for grade in list(grades):
            try:
                if isinstance(int(grade), int):
                    grades_list.append(grade)
            except ValueError:
                pass
    grade_choices = [(i, i) for i in grades_list]
    return grade_choices

def getTopicChoices(subject,board,grade):
    # print "subject,board,grade",subject,board,grade
    course_id=Course.objects.filter(board_name=board, subject=subject,grade__icontains=grade).values_list("id",flat=True).distinct()[:1]
    topics = Topic.objects.filter(course_id=course_id).values_list("id", "title").order_by("id")
    topics_choices = [(i[0], i[1]) for i in topics]
    return topics_choices

def getSubTopicChoices(topic_id):
    subtopics = SubTopics.objects.filter(topic_id=topic_id).values_list("id", "name").order_by("id")
    subtopics_choices = [(i[0], i[1]) for i in subtopics]
    return subtopics_choices


def getAddQuestionDropdowns(request,type_name):
	if request.method == "GET":
		if type_name == "getsubjects":
			board = request.GET.get("board", "")
			subject_choices = getSubjectChoices(board)
			return HttpResponse(json.dumps(subject_choices), content_type='application/json')
		elif type_name == "getgrades":
			board = request.GET.get("board", "")
			subject = request.GET.get("subject", "")
			grade_choices = getGradeChoices(board, subject)
			return HttpResponse(json.dumps(grade_choices), content_type='application/json')
		elif type_name == "topics":
			subject = request.GET.get("subject", "")
			board=request.GET.get("board", "")
			grade=request.GET.get("grade", "")
			topics_choices = getTopicChoices(subject,board,grade)
			return HttpResponse(json.dumps(topics_choices), content_type='application/json')
		elif type_name == "subtopics":
			topic = int(request.GET.get("topic", ""))
			subtopics_choices = getSubTopicChoices(topic)
			return HttpResponse(json.dumps(subtopics_choices), content_type='application/json')


def getAssesment(request):
    is_eligible = True
    userprofile_id =  UserProfile.objects.values_list("id",flat = True).filter(user_id = request.user.id)
    role_id = RolePreference.objects.filter(userprofile_id = userprofile_id,role_id =3,role_status = "Active",role_outcome = "Recommended")
    if request.user.is_superuser:
        board_list = Center.objects.filter(status="Active").values_list("board", flat=True).distinct().order_by("board")
    elif role_id:
        board_list = Center.objects.filter(status="Active").values_list("board", flat=True).distinct().order_by("board")
    elif  has_role(request.user.userprofile, "Teacher") or has_pref_role(request.user.userprofile, "Teacher"):

        board_list = Offering.objects.filter(active_teacher__id=request.user.id).values_list("center__board",flat=True).distinct()
        if len(board_list) == 0:
                is_eligible = False
    else:
        board_list = Center.objects.filter(Q(admin=request.user) | Q(assistant=request.user) 
                                    | Q(field_coordinator_id=request.user.id) | Q(delivery_coordinator_id=request.user.id)
                                    ).filter(status="Active").values_list("board", flat=True).distinct().order_by("board")
    if len(board_list) == 0:
        try:
            board_list = Center.objects.filter(Q(funding_partner_id=request.user.partner_set.all()[0].id) \
                                    | Q(delivery_partner=request.user.partner_set.all()[0].id) | Q(orgunit_partner=request.user.partner_set.all()[0].id) \
                                    ).filter(status="Active").values_list("board", flat=True).distinct().order_by("board")
        except IndexError:
            board_list = ''
    # board_list = []
    # center_list = []
    # offering_list = []
    # for center in centers:
    #     board_list.append(center.board)
    #     center_list.append(center.name)
    # board_list = list(set(board_list))
    # center_list = list(set(center_list))
    board_list = sorted(board_list)
    context ={
        'is_eligible' : is_eligible,
        'board_list' : board_list,
    }

    # topics = Topic.objects.filter(course_id_id = 20)
    # subtopic_list = SubTopics.objects.filter(topic_id = 50)
    return render(request, 'assesment_builder.html',context)

def get_centers_by_state(request):
	board = request.GET.get('state')
	center_json = {}
	userprofile_id =  UserProfile.objects.values_list("id",flat = True).filter(user_id = request.user.id)
	role_id = RolePreference.objects.filter(userprofile_id = userprofile_id,role_id =3,role_status = "Active",role_outcome = "Recommended")

	if request.user.is_superuser:
		center_list = Center.objects.filter(status="Active",board = board).values("name", "id").distinct().order_by("board")
	elif role_id:
	    center_list = Center.objects.filter(status="Active",board = board).values("name", "id").distinct().order_by("board")
	elif  has_role(request.user.userprofile, "Teacher") or has_pref_role(request.user.userprofile, "Teacher"):

		center_list = Offering.objects.filter(active_teacher__id=request.user.id).values_list("center__name",flat=True).distinct()
        
	else:
		center_list = Center.objects.filter(Q(admin=request.user) | Q(assistant=request.user) 
                                    | Q(field_coordinator_id=request.user.id) | Q(delivery_coordinator_id=request.user.id)
                                    ).filter(status="Active",board = board).values("name", "id").distinct().order_by("board")
	if len(center_list) == 0:
		try:
			center_list = Center.objects.filter(Q(funding_partner_id=request.user.partner_set.all()[0].id) \
                                    | Q(delivery_partner=request.user.partner_set.all()[0].id) | Q(orgunit_partner=request.user.partner_set.all()[0].id) \
                                    ).filter(status="Active",board = board).values("name", "id").distinct().order_by("board")
		except IndexError:
			center_list = ''
	try:
		center_list =Center.objects.filter(status="Active",name__in = center_list).values("name","id").distinct().order_by("board")
	except:
		pass



	# if board:
	# 	centers_list = Center.objects.filter(status="Active",board = board).values("name","id").distinct().order_by("board")
	print("#####################33",center_list)
	for center in center_list:
		center_json[center['id']] = center['name']
	return HttpResponse(json.dumps({'center_json':center_json}), mimetype='application/json')
def get_offering_by_center(request):
	center_id = request.GET.get('center')
	center = Center.objects.get(id = int(center_id))
	academic_year_id  = Ayfy.objects.filter(board = center.board).order_by('-id')[:1]
	Teacher = Offering.objects.filter(active_teacher__id = request.user.id)
	userprofile_id =  UserProfile.objects.values_list("id",flat = True).filter(user_id = request.user.id)
	role_id = RolePreference.objects.filter(userprofile_id = userprofile_id,role_id =3,role_status = "Active",role_outcome = "Recommended")
	offering_list = []
	offering_json = {}
	if role_id or request.user.is_superuser:
	    offering_list = Offering.objects.filter(academic_year_id = academic_year_id,center_id = center_id)
	elif Teacher:
	    offering_list = Offering.objects.filter(academic_year_id = academic_year_id,center_id = center_id,active_teacher__id = request.user.id)
	else:
	    offering_list = Offering.objects.filter(academic_year_id = academic_year_id,center_id = center_id)
	print("@@@@@@@@@@@@@@@@@@@",offering_list)
	for offering in offering_list:
	    offering_json[offering.id] = [str(offering.course.grade)+'th '+str(offering.course.subject)
                    +' '+str(offering.start_date.strftime("%Y-%m-%d"))+"-"+
	                str(offering.end_date.strftime("%Y-%m-%d")),offering.course.id]
	return HttpResponse(json.dumps({'offering_json':offering_json}), mimetype='application/json')
def get_topic_and_sub_topic(request):
	course_id = request.GET.get('course_id')
	topic_id = request.GET.get('topic_id')
	if course_id:
		course = Course.objects.get(id = int(course_id))
		topics_choices = getTopicChoices(course.subject,course.board_name,course.grade)
		return HttpResponse(json.dumps(topics_choices), content_type='application/json')
	elif topic_id:
		temp = re.findall(r'\d+', topic_id)
		topicList = list(map(int, temp))
		print "topic list",topicList
		subtopics_choices = []
		for i in range(0,len(topicList)):
			subtopics_choices += getSubTopicChoices(int(topicList[i]))
		print "subtopicsss",subtopics_choices
		return HttpResponse(json.dumps(subtopics_choices), content_type='application/json')
def get_question_list(request):
	topic_id = request.GET.getlist('topic_id[]')
	print "topic ...........",topic_id
	sub_topics_id = request.GET.getlist('sub_topic_id[]')
	page = request.GET.get('page', 1)
	filter = {}
	# if sub_topics_id:
	# 	for i in sub_topics_id:
	# 	    filter['subtopic_id'].append(i)
	# if topic_id:
	# 	for i in topic_id:
	# 	    filter['topic_id'].append(i)
	# print "filterrrr",filter
	question_json = {}
	board = request.GET.get('state')
	course_id = request.GET.get('course_id')
	if board:
		center = Center.objects.filter(board = board).distinct().order_by('board')[:1][0]
		if center:
			filter['board'] = str(center.board)

	if course_id:
		course = Course.objects.get(id = int(course_id))
		if course:
			filter['subject'] = str(course.subject)
			filter['grade'] = str(course.grade)
	if len(sub_topics_id)==0:
		questionlist = Question.objects.filter(**filter).filter(topic_id__in =topic_id)
	else:
		questionlist = Question.objects.filter(**filter).filter(topic_id__in =topic_id).filter(subtopic_id__in =sub_topics_id)

	paginator = Paginator(questionlist, 50)
	try:
		questionlist = paginator.page(page)
	except PageNotAnInteger:
		questionlist = questionlist.page(1)
	except EmptyPage:
		questionlist = paginator.page(paginator.num_pages)

	for qsn in questionlist:
		subtopic_name = ''
		if qsn.subtopic:
			subtopic_name = qsn.subtopic.name
		data = {'topic':qsn.topic.title,'sub_topic':subtopic_name,'question':qsn.question,'type':qsn.question_type,'level':qsn.difficulty_level,'marks':qsn.marks, 'choices':qsn.answer_question, 'subject':qsn.subject, 'grade':qsn.grade, 'correctAnswer':qsn.correct_answer,"image":str(qsn.question_image)}
		question_json[qsn.id] = data
	return HttpResponse(json.dumps({'questionlist':question_json,'prev': questionlist.previous_page_number(),
			 'next': questionlist.next_page_number(), 'current': questionlist.number,
			 'total': questionlist.paginator.num_pages, 'count': len(questionlist)}), content_type='application/json')


def approve_reject_question(request):
    try:
        if request.method == 'GET':
            question_id = request.GET.get('id', '')
            is_approve = request.GET.get('approve', '')
            print request.GET
            if question_id == '' and is_approve:
                resp = {'data' : {'message' : 'Invalid Request'}}
                return HttpResponseRedirect("/questions")
            question = Question.objects.filter(id=question_id)
            if question.count():
                approve_status = 'Approved'
                if is_approve == 0 or is_approve == '0':
                    approve_status = 'Rejected'
                question.update(status= approve_status)
                return HttpResponseRedirect('/questions/')
            else:
                return HttpResponseRedirect('/questions/')
        else:
            return HttpResponseRedirect('/questions/')
    except Exception as e:
        pass
        # print "Error reason = ", e
        # print "Error at line no = ", traceback.format_exc()
    return HttpResponseRedirect('/questions/')
def question_instruction(request):
    question_type = request.POST.get('question_type')
    try:
        question_instruction = QuestionType.objects.values_list('instructions',flat=True).filter(name = question_type)[0]
    except:
        question_instruction = "NA"
    response = {}
    response['instructions'] = question_instruction
    return HttpResponse(simplejson.dumps(response), mimetype='application/json')



@login_required()
def bulk_upload_quiz_view(request):
    return render(request, "bulk_upload_quiz_view.html", {})


def validateandGetQuizExcelHeaders(sheetObj,maxCol):
    try:
        expectedFieldsArray = ['BoardCode', 'ContentDetailId', 'CourseId', 'LanguageCode', 'TopicId', 'SubtopicId',
                               'complexity',
                               'priority', 'points', 'QuestionFormat', 'formatType', 'questionStem', 'text', 'imageUrl',
                               'option1',
                               'option2', 'option3', 'option4', 'option5', 'option6', 'correctOptions',
                               'correctOptionSequence',
                               'columnBOption1', 'columnBOption2', 'columnBOption3', 'columnBOption4', 'category1',
                               'category2',
                               'categoryItems', 'categoryItemParent', 'Learning Outcome']

        fieldsArray = []
        for i in range(1, maxCol + 1):
            cellObj = sheetObj.cell(row=1, column=i)
            if cellObj.value != None:
                fieldName = expectedFieldsArray[i-1]
                if fieldName != cellObj.value:
                    return None
                fieldsArray.append(cellObj.value)

        return fieldsArray
    except Exception as e:
        print("validateandGetQuizExcelHeaders Exceptions",e)
        return None



def getResultObject(status,message,object,objectType):
    data =  {
        "status":status,
        "message":message
    }
    if objectType:
        data[objectType] = object
    return data

def getQuestionComplexityFromData(compStr):
    compStr = str(compStr)
    compVal = 1
    if compStr == "Medium":
        compVal = 2
    elif compStr == "Advanced":
        compVal = 3
    return compVal

def createQuestionComponentRec(optionVal,questionObj,hasCorrectAns,subType,matchCompId):
    compObj = Question_Component.objects.create(
        text=optionVal,
        question=questionObj,
        subtype=subType,
        is_answer=hasCorrectAns,
        created_by_id=settings.SYSTEM_USER_ID_AUTH,
        updated_by_id=settings.SYSTEM_USER_ID_AUTH,
        matching_component_id = matchCompId
    )
    compObj.save()
    return compObj

def createMCQQuestionComponents(dataObject,questionObj,qFormat):
    try:
        subType = '1'
        corValList = []
        if qFormat == "mcq" or qFormat == "mmcq":
            corVals = dataObject.get("correctOptions")
            corVals = str(corVals)
            corValList = corVals.split(",")
        else:
            subType = '6'
        optionidsArray = []
        for i in range(0,6):
            optionColStr = "option"+str((i+1))
            optionVal = dataObject.get(optionColStr)
            if optionVal and optionVal != '':
                # createOptionItem
                hasCorrectAns= False
                if optionColStr in corValList:
                    hasCorrectAns = True
                compObj = createQuestionComponentRec(optionVal,questionObj,hasCorrectAns,subType,None)
                optionidsArray.append(compObj.id)

        resultObj = getResultObject(1,None, None, None)
        if qFormat == "orderBy":
            actualOrderArr = []
            corVals = dataObject.get("correctOptionSequence")
            corVals = str(corVals)
            seqList = corVals.split(",")
            for item in seqList:
                indexVal = item[len(item) - 1]
                indexVal = int(indexVal)
                if len(optionidsArray) > (indexVal-1):
                    actualOrderArr.append(optionidsArray[indexVal-1])

            resultObj["actualSequence"] = genUtility.getStringFromIdArray(actualOrderArr)
        return resultObj
    except Exception as e:
        print("createMCQQuestion Exceptions", e)
        #traceback.print_exc()
        logService.logException("createMCQQuestionComponents Exceptions", e.message)
        return getResultObject(0,e.message,None,None)

def createMatchTheFollowingQuestionComponents(dataObject,questionObj,qFormat):
    try:
        optionidsArray = []
        corVals = dataObject.get("correctOptionSequence")
        corVals = str(corVals)
        seqList = corVals.split(",")

        for i in range(0, 6):
            optionColStr = "columnBOption" + str((i + 1))
            optionVal = dataObject.get(optionColStr)
            if optionVal and optionVal != '':
                compObj = createQuestionComponentRec(optionVal, questionObj, False, '3',None)
                optionidsArray.append(compObj.id)


        for i in range(0,6):
            matchIngOptionid = 0
            optionColStr = "option"+str((i+1))
            optionVal = dataObject.get(optionColStr)
            if optionVal and optionVal != '':
                correctColumnName = seqList[i]
                indexVal = correctColumnName[len(correctColumnName) - 1]
                indexVal = int(indexVal)
                if len(optionidsArray) > (indexVal - 1):
                    matchIngOptionid = optionidsArray[indexVal - 1]
                createQuestionComponentRec(optionVal,questionObj,False,'2',matchIngOptionid)
                

        resultObj = getResultObject(1, None, None, None)
        return resultObj
    except Exception as e:
        print("createMatchTheFollowingQuestion Exceptions", e)
        #traceback.print_exc()
        logService.logException("createMatchTheFollowingQuestionComponents Exceptions", e.message)
        return getResultObject(0,e.message,None,None)


def createCategoriseQuestionComponents(dataObject,questionObj,qFormat):
    try:
        subType = '4'
        category1 = dataObject.get("category1")
        category2 = dataObject.get("category2")

        categoryItemStr = dataObject.get("categoryItems")
        categoryItems = categoryItemStr.split(",")

        categoryMapStr = dataObject.get("categoryItemParent")
        categoryMapStr = str(categoryMapStr)
        categoryMapItems = categoryMapStr.split(",")
        if len(categoryItems) != len(categoryMapItems):
            return getResultObject(0, "Category item is not matching with categoryItemParent", None, None)

        catObj1 = createQuestionComponentRec(category1, questionObj, False, subType, None)
        catObj2 = createQuestionComponentRec(category2, questionObj, False, subType, None)

        for i in range(0, len(categoryItems)):
            optionVal = categoryItems[i]
            if optionVal and optionVal != '':
                categoryMap = categoryMapItems[i]
                parentCat = catObj1
                if categoryMap == "category2":
                    parentCat = catObj2

                createQuestionComponentRec(optionVal, questionObj, False, '5',parentCat.id)

        resultObj = getResultObject(1, None, None, None)
        return resultObj
    except Exception as e:
        print("createCategoriseQuestionComponents Exceptions", e)
        #traceback.print_exc()
        logService.logException("createCategoriseQuestionComponents Exceptions", e.message)
        return getResultObject(0,e.message,None,None)


def createQuestion(dataObject,questionSet,subtopicObj,topicId):
    try:
        title = dataObject.get("questionStem")
        text = dataObject.get("text")
        if text is None:
            text = ''
        complexity = getQuestionComplexityFromData(dataObject.get("complexity"))
        points = dataObject.get("points")
        priority = dataObject.get("priority")
        learning_outcome = dataObject.get("Learning Outcome")
        qFormat = dataObject.get("QuestionFormat")
        qFormat = str(qFormat)
        questionType = Question_Type.objects.get(code=qFormat)

        questObj = Question.objects.create(
            title = title,
            text = text,
            questionset_id=questionSet.id,
            topic_id = topicId,
            subtopic_id = subtopicObj.id,
            points=points,
            sequence=priority,
            created_by_id=settings.SYSTEM_USER_ID_AUTH,
            updated_by_id=settings.SYSTEM_USER_ID_AUTH,
            learning_outcome =learning_outcome,
            complexity = complexity,
            type = questionType,
            type_code = questionType.code
        )
        questObj.save()

        resultSet = None
        if qFormat == "mcq":
            resultSet = createMCQQuestionComponents(dataObject,questObj,qFormat)
        elif qFormat == "mmcq":
            resultSet = createMCQQuestionComponents(dataObject, questObj, qFormat)
        elif qFormat == "matchTheFollowing":
            resultSet = createMatchTheFollowingQuestionComponents(dataObject,questObj,qFormat)
        elif qFormat == "orderBy":
            resultSet = createMCQQuestionComponents(dataObject, questObj, qFormat)
            questObj.actual_sequence_string = resultSet["actualSequence"]
            questObj.save()
        elif qFormat == "categorise":
            resultSet = createCategoriseQuestionComponents(dataObject, questObj, qFormat)

        if resultSet["status"] == 0:
            return resultSet
        return getResultObject(1,None,None,None)
    except Exception as e:
        print("createQuestion Exceptions", e)
        logService.logException("createQuestion Exceptions", e.message)
        return getResultObject(0,e.message,None,None)


def createQuestionItem(itemObject,question):
    try:
        pass
    except Exception as e:
        print("createQuestionItem", e)
        return None


def getSubtopicOfQuiz(courseId,topicId):
    try:
        subtopicObjList = SubTopics.objects.filter(topic_id=topicId,type='2')
        if subtopicObjList and len(subtopicObjList) > 0:
            return subtopicObjList[0]
        return None
    except Exception as e:
        print("getSubtopicOfQuiz", e)
        return None

def getQuestionSetForId(topicId,subtopicId):
    try:
        objList = Question_Set.objects.filter(topic_id=topicId,subtopic_id=subtopicId)
        if objList and len(objList) > 0:
            return objList[0]
        return None
    except Exception as e:
        print("getQuestionSetForId", e)
        return None


def insertSubtopicOfQuiz(courseId,topicId):
    try:
        dateObj = genUtility.getCurrentTime()
        subtopic = SubTopics.objects.create(
            topic_id=topicId,
            type='2',
            name='Quiz',
            created_date=dateObj,
            updated_date = dateObj,
            created_by_id = settings.SYSTEM_USER_ID_AUTH,
            updated_by_id=settings.SYSTEM_USER_ID_AUTH,
            status = 'Active'
        )
        subtopic.save()
        return subtopic
    except Exception as e:
        print("insertSubtopicOfQuiz", e)
        logService.logException("insertSubtopicOfQuiz Exceptions", e.message)
        return None

def insertQuestionSetForSubtopic(courseId,topicId,subtopic,contentDetail):
    try:
        qSet = Question_Set.objects.create(
            topic_id=topicId,
            subtopic_id=subtopic.id,
            name='Quiz',
            created_by_id = settings.SYSTEM_USER_ID_AUTH,
            updated_by_id=settings.SYSTEM_USER_ID_AUTH,
            type = 1,
            content_detail = contentDetail
        )
        qSet.save()
        return qSet
    except Exception as e:
        print("insertQuestionSetForSubtopic", e)
        logService.logException("insertQuestionSetForSubtopic Exceptions", e.message)
        return None

def insertContentDetailForSubtopic(courseId,topicId,subtopic):
    try:
        workstream = WorkStreamType.objects.get(code="quiz")
        contentType = ContentTypeMaster.objects.get(code="quiz-evd")


        cdObj = ContentDetail.objects.create(
            topic_id=topicId,
            subtopic_id=subtopic.id,
            name='Quiz',
            created_by_id = settings.SYSTEM_USER_ID_AUTH,
            updated_by_id=settings.SYSTEM_USER_ID_AUTH,
            workstream_type = workstream,
            status = "approved",
            is_primary = True,
            content_type = contentType
        )
        cdObj.save()
        return cdObj
    except Exception as e:
        print("insertContentDetailForSubtopic", e)
        logService.logException("insertContentDetailForSubtopic Exceptions", e.message)
        return None

def insertQuizData(dataArray):
    try:
        #Check if topic exist
        #For topic, check if subtopic of type quiz exist, if not create
        #Create question for the subtopic with name
        #create Content detail record if not exist
        #check question type, extra the Columns accordingly
        #create question
        #create question component
        #update question
        rowNumber = 1
        for eachDataObj in dataArray:
            rowNumber = rowNumber + 1
            courseIdStr = eachDataObj.get("CourseId")
            topicIdStr = eachDataObj.get("TopicId")
            if courseIdStr and topicIdStr:
                topicId = int(topicIdStr)
                courseId = int(courseIdStr)
                subtopic = getSubtopicOfQuiz(courseId,topicId)
                questionSet = None
                if subtopic is None:
                    #create subtopic object
                    subtopic = insertSubtopicOfQuiz(courseId, topicId)
                    if subtopic is None:
                        print("Subtopic exception")
                        return getResultObject(0,"Subtopic exception for row number "+ str(rowNumber),None,None)
                    else:
                        contentDetail = insertContentDetailForSubtopic(courseId, topicId,subtopic)
                        questionSet = insertQuestionSetForSubtopic(courseId, topicId,subtopic,contentDetail)
                else:
                    questionSet = getQuestionSetForId(topicId,subtopic.id)
                    if questionSet is None:
                        return getResultObject(0,"Question set issue for row number "+ str(rowNumber),None,None)

                resultData = createQuestion(eachDataObj, questionSet, subtopic, topicId)
                if resultData["status"] == 1:
                    continue
                else:
                    errMsg = resultData["message"]
                    if errMsg:
                        resultData["message"] = errMsg + " row number "+str(rowNumber)
                    else:
                        resultData["message"] = "Error in row number "+ str(rowNumber)

                    return resultData
            else:
                return getResultObject(0,"Topic id missing for row number "+ str(rowNumber),None,None)

        return getResultObject(1,None,None,None)

    except Exception as e:
        print("insertQuizData Exceptions",e)
        logService.logException("insertQuizData Exceptions", e.message)
        return  getResultObject(0,e.message,None,None)




@login_required()
def bulk_upload_quiz_action(request):
    input_excel = request.FILES['file']
    wb_obj = load_workbook(input_excel)
    sheetObj = wb_obj.active
    maxCol = sheetObj.max_column
    maxRow = sheetObj.max_row

    fieldsArray = validateandGetQuizExcelHeaders(sheetObj,maxCol)
    dataArray = []
    for rowCounter in range(2,maxRow):
        datadict = {}
        colCounter = 1
        for eachField in fieldsArray:
            cellObj = sheetObj.cell(row=rowCounter, column=colCounter)
            colCounter=colCounter + 1
            #Add validations here
            if cellObj.value != None:
                datadict[eachField] = cellObj.value
        if len(datadict) <= 0:
            break
        dataArray.append(datadict)

    #print("dataArray",dataArray)
    totalLen = len(dataArray)
    resultData = insertQuizData(dataArray)
    if resultData["status"] == 1:
        return render(request, "bulk_upload_quiz_view.html", {"message": "Successfully uploaded and processed the records. Total number of questions " + str(totalLen)})
    else:
        return render(request, "bulk_upload_quiz_view.html", {"rejectedData": resultData["message"]})
    #return HttpResponseRedirect("/questions/upload/quiz/view")
